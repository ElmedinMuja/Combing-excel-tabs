import pandas as pd
import openpyxl
import os
import numpy as np
import shutil

# ## create seperate blank file with new tab to copy data to
# ## new_file_name = Baseline Ready + filename
# ## tabname = Baseline_data
# ## requres only original file path as arg
# def create_baselineReady_file(filename):
#
#     ## create new filename
#     basename = os.path.dirname(filename) ## file path without filename
#     name_of_file = os.path.basename(filename) ## filename string
#     baseline_ready_filename = basename+'\\'+'Baseline Ready ' + name_of_file
#     wb = openpyxl.Workbook() ## open blank wb
#
#     ## search if file already exists in folder
#     for files in os.listdir(basename):
#         if 'Baseline Ready ' in files: # if 'Baseline Ready' in a filename, then
#             pass # end loop
#             print('Baseline Ready File already exists.') # notify user
#             return baseline_ready_filename
#
#
#     else: # if 'Baseline Ready' not found in a filename, then
#         wb.save(baseline_ready_filename) # save the open workbook as a new file
#
#         ## open and save a new workbook with new name for baseline ready
#
#         ws = wb['Sheet'] ## set var to sheet name
#         ws.title = 'Baseline_data' ## rename existing sheet to new name
#         # wb.create_sheet('Baseline_data')
#         wb.save(baseline_ready_filename)
#
#         ## headers list to add to the file
#         ## headers list to add to the file
#         Header_list2 = [ # 'Platform/Partner',
#             'media_audience_dmp_name', #
#             # 'Start Date (Media Audience)',
#             # 'End Date (Media Audience)',
#             'mobile_app_ad_group_id', #
#             'desktop_mobile_web_ad_group_id', #
#             # 'Media Audience Segment ID ',
#             'media_audience_tactic_id', #
#             'creative_audience_dmp_name', #
#             # 'Start Date (Creative Audience)',
#             # 'End Date (Creative Audience)',
#             'Creative Audience Segment ID', #
#             'creative_audience_tactic_id', #
#             'Variant',#
#             'Creative_Language',#
#             'Creative_Theme',#
#             'Personalization_1', #
#             'Personalization_2', #
#             'Personalization_3', #
#             'Claim', #
#             'CTA', #
#             # 'Landing Page',
#             'Retailer', #
#             # 'Ad Size',
#             # 'Ad Format',
#             # 'Modeled Version Name',
#             'Creative_Version_Name',#
#             # 'Creative Version Name Character Count',
#             # 'Rotation',
#             'click_through_url_with_webtrends_if_applicable', #
#             # 'Link to Creative',
#             # 'Notes',
#             'Brand', #
#             # 'Creative Version Preview Link',
#             'creative_audience_brand_friendly_name', #
#             'media_audience_brand_friendly_name' #
#         ]
#         ## set sheetname; ws obj
#         ws = wb['Baseline_data']
#
#         ## iterate through the columns on row 1 and add the headers in the order they are in the headers_list2
#         for column in range(1,len(Header_list2)):
#             coordinate = ws.cell(row=1,column=column).coordinate
#             ws[coordinate] = Header_list2[column-1]
#
#         ## save and close workbook, a must to preserve changes and not corrupt file
#         wb.save(baseline_ready_filename)
#         wb.close()
#         return baseline_ready_filename
#         print(f"New file created @: {baseline_ready_filename}")
#
# fp = r'data_file.xlsx'
# create_baselineReady_file(r'C:\Users\Dini\Python\py\workFunctions\DQ Scripts\KC opt v2\data_file.xlsx')
'''
GOAL:
    TO COMBINE MULTIPLE TABS INTO ONE FILE. 

STEPS:
    1 -     USER WILL PASS A FILE WITH MULTI TABS THEY WANT TO COMBINE AND A NEW EXCEL FILE WITH THE HADERS ALREADY ADDED [must be csv]
    2 -     USER WILL INDICATE WHICH TAB THEY WANT TO ADD TO THE BLANK FILE; they will have to do this multiple times for multiple tabs
    3 -     USER WILL THEN MAP THE COLUMNS BETWEEN THE TWO FILES
    4 -     SCRIPT WILL GET COLUMN DATA FROM BOTH, COMBINE AND ADD THEM TO AN EMPTY DF
    5 -     ONCE ALL COLUMNS ARE ITERATED, SCRIPT WILL SAVE THE NEW DATAFRAME TO THE NAME OF THE BLANK FILE USER PASSED ON
    
NOTES:
    Currently, script only does one tab at a time, so to do multiple tabs, itll have to be run multi times
    but, the data from the new tab will always be appended to the bsl file

TO ADD:
    1 -     HAVE THE SCRIPT ASK FOR ALL TABS USER WANTS TO COMBINE, THEN LOOP THROUGH ALL THOSE
        a - USER WILL NEED TO MAP FOR EACH TAB, EVERYTIME ONE IS COMPLETE, NEW UI WILL OPEN UP AND HAVE THEM MAP IT
    
    2 -     WHEN USER PASSES DATAFILE, CREATE A COPY AND STORE IT IN FOLDER THEYRE WORKING OUT OF   
    3 - X -     REMOVE ROWS OF WHITESPACE


ISSUSE:
1 -     SWAPPING THE TAB NAMES IN TABS LIST WORKS DIFFERENT
        USING : tabs_list =['OBP TTD OLA','OBP DV360 OLA'], YOU GET BOTH DATA SETS IN ONE TAB
        BUT SWAPPING ONLY ALLOWS FOR THE DV360

'''



fullfilepath = r'C:\Users\Dini\Python\py\workFunctions\DQ Scripts\KC opt v2\data_file.xlsx'
fn = 'data_file.xlsx'
fn_merge = r'C:\Users\Dini\Python\py\workFunctions\DQ Scripts\KC opt v2\Baseline Ready data_file.csv'

## if they dont already exist, create copy of bsl file

## baseline file
directory = os.path.dirname(fn_merge)
base_filename = os.path.basename(fn_merge)
newfile = directory + '\\[COPY] Baseline Ready data_file.csv'
if os.path.exists(newfile) == True:
    pass
elif os.path.exists(newfile) == False:
    shutil.copy(fn_merge,newfile)




#user_input_tab = input('what tab are we adding to the file? \n')
tabs_list =['OBP TTD OLA','OBP DV360 OLA']

user_input_tab = 'OBP DV360 OLA'
user_input_column_og = 'Variant'
user_input_column_new = 'Variant'

datafile_column_list = [
    'Personalization 1',
    'Personalization 2',
    'Personalization 3',

                        ]
bslfile_column_list = [
    'Personalization_1',
    'Personalization_2',
    'Personalization_3',
]

# check to see if the tab we're looking at in the excel file has a first row that != the headers we need
for tabs in tabs_list:
    df_datafile= pd.read_excel(fn,tabs)
    if df_datafile.columns.values[0] == 'FCB':
        print(f"Please ensure the first row for tab: \n{tabs} are the headers of the sheet.")
        sys.exit()
    else:
        pass



# # prnt all columns of files
# for items in df_bsl.columns.values: print(items) # bsl file
# print('\n')
# for items in df_datafile.columns.values: print(items) # datafile
df = pd.DataFrame()
for tabs in tabs_list:
    print('\n',tabs)
    # create empty dataframe to store the data


    # create dataframes for the file we're taking data from and file were importing data into
    df_datafile = pd.read_excel(fn,tabs) # file with data
    df_bsl = pd.read_csv(fn_merge) # BSL file that will be uploaded





    # start empty df where we will append series into
    # loop through the headers from headerlist
    for index in range(0,len(datafile_column_list)):

        # variable for the column names
        columnName = datafile_column_list[index] ## data file column name
        columnName_bsl = bslfile_column_list[index] ## bsl file column name
        print(columnName+ '  |  ' +columnName_bsl)

        print(df_datafile[columnName])
        # get column data and turn into a series from datafile
        series_1 = pd.Series(df_datafile[columnName])
        print(series_1)
        # get column data and turn into a series from bslfile
        series_2 = pd.Series(df_bsl[columnName_bsl])

    # combine both series into one and give the column the name of the data file [the correct header bsl will take]
        series_3 = pd.concat([series_1,series_2],axis=0)


        # remove the blanks/blank rows in the file;
        # vendors usually drag the dropdown so while they look empty, pandas detects them as cells with values
        series_3.dropna(axis=0,
                  how='all',
                  inplace=True)


#         # add combined series to the df
#         df[columnName_bsl] = series_3
#
# df.to_csv('test.csv',index=False)


'''
adding a 3rd tab: Impressa TTD OLA causes it to return keyerror. Is it because of the values in the header thats diff?
or is it the script

'''
